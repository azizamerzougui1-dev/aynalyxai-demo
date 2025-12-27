# AynalyxAI Demo - Streamlit Cloud Version
# Sample data only - no file upload

import streamlit as st
import pandas as pd
import numpy as np
from sklearn.ensemble import IsolationForest
from sklearn.preprocessing import StandardScaler
from io import BytesIO
from openpyxl.styles import PatternFill, Font, Border, Side

st.set_page_config(page_title="AynalyxAI Demo", page_icon="🔬", layout="centered", initial_sidebar_state="collapsed")

st.markdown("""
<style>
    [data-testid="stSidebar"] { min-width: 320px; max-width: 400px; }
    .main-header { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 2rem; border-radius: 16px; color: white; text-align: center; margin-bottom: 1.5rem; }
    .main-header h1 { margin: 0; font-size: 2.5rem; }
    .main-header p { margin: 0.5rem 0 0 0; opacity: 0.95; font-size: 1.1rem; }
    .demo-badge { background: #fef3c7; color: #92400e; padding: 0.75rem 1.25rem; border-radius: 10px; font-weight: 600; display: block; text-align: center; margin-bottom: 1.5rem; border: 2px solid #f59e0b; }
    .feature-box { background: linear-gradient(135deg, #f0f9ff 0%, #e0f2fe 100%); padding: 1.5rem; border-radius: 12px; margin-bottom: 1.5rem; border-left: 4px solid #0ea5e9; }
    .feature-box h3 { color: #0369a1; margin: 0 0 0.5rem 0; }
    .feature-box p { color: #475569; margin: 0; line-height: 1.6; }
    .stat-card { background: white; padding: 1rem; border-radius: 10px; text-align: center; box-shadow: 0 2px 8px rgba(0,0,0,0.08); border-left: 4px solid; }
    .stat-critical { border-color: #dc2626; }
    .stat-high { border-color: #ef4444; }
    .stat-medium { border-color: #f59e0b; }
    .stat-low { border-color: #22c55e; }
    .sidebar-title { font-size: 1.3rem; font-weight: bold; color: #1e293b; margin-bottom: 1rem; padding-bottom: 0.5rem; border-bottom: 2px solid #667eea; }
    .how-it-works { background: #f8fafc; padding: 1.25rem; border-radius: 10px; margin: 1rem 0; border: 1px solid #e2e8f0; }
    .cta-box { background: linear-gradient(135deg, #667eea15 0%, #764ba215 100%); padding: 1.5rem; border-radius: 12px; text-align: center; border: 2px solid #667eea; margin-top: 2rem; }
    .cta-box h3 { color: #667eea; margin-bottom: 0.5rem; }
    .cta-button { display: inline-block; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white !important; padding: 0.75rem 1.5rem; border-radius: 25px; text-decoration: none; font-weight: bold; margin-top: 1rem; }
    .advantage-grid { display: grid; grid-template-columns: repeat(3, 1fr); gap: 1rem; margin: 1rem 0; }
    .advantage-item { background: white; padding: 1rem; border-radius: 10px; text-align: center; box-shadow: 0 2px 6px rgba(0,0,0,0.06); }
    .advantage-item .icon { font-size: 2rem; margin-bottom: 0.5rem; }
    .advantage-item h4 { color: #334155; margin: 0 0 0.25rem 0; font-size: 0.95rem; }
    .advantage-item p { color: #64748b; margin: 0; font-size: 0.85rem; }
    
    /* ========== MOBILE RESPONSIVE CSS ========== */
    @media (max-width: 768px) {
        .main .block-container { padding: 1rem 0.5rem !important; }
        .main-header { padding: 1rem !important; border-radius: 8px !important; }
        .main-header h1 { font-size: 1.5rem !important; }
        .main-header p { font-size: 0.9rem !important; }
        [data-testid="column"] { width: 100% !important; flex: 1 1 100% !important; }
        .stat-card { padding: 0.75rem !important; margin-bottom: 0.5rem !important; }
        .advantage-grid { grid-template-columns: 1fr !important; gap: 0.75rem !important; }
        .feature-box { padding: 1rem !important; }
        .cta-box { padding: 1rem !important; }
        .stDataFrame { max-height: 350px !important; overflow-x: auto !important; -webkit-overflow-scrolling: touch !important; }
        .stButton button { padding: 0.5rem 1rem !important; font-size: 0.9rem !important; }
        .demo-badge { padding: 0.5rem 0.75rem !important; font-size: 0.85rem !important; }
    }
    @media (max-width: 480px) {
        .main-header h1 { font-size: 1.3rem !important; }
        .advantage-item { padding: 0.75rem !important; }
        .advantage-item .icon { font-size: 1.5rem !important; }
        .advantage-item h4 { font-size: 0.85rem !important; }
        .advantage-item p { font-size: 0.75rem !important; }
    }
</style>
""", unsafe_allow_html=True)

with st.sidebar:
    st.markdown('<p class="sidebar-title">🌐 Language / Langue</p>', unsafe_allow_html=True)
    lang = st.selectbox("", ["English", "Francais"], label_visibility="collapsed")
    is_fr = lang == "Francais"

t = {
    "title": "AynalyxAI Demo" if not is_fr else "Demo AynalyxAI",
    "subtitle": "AI-Powered Financial Anomaly Detection" if not is_fr else "Detection d Anomalies Financieres par IA",
    "demo_notice": "🎯 FREE DEMO - Test our AI with sample financial data!" if not is_fr else "🎯 DEMO GRATUITE - Testez notre IA!",
    "sample_title": "📊 Select Sample Data" if not is_fr else "📊 Choisir des Donnees",
    "sample_invoices": "📄 Invoices" if not is_fr else "📄 Factures",
    "sample_expenses": "💰 Expenses" if not is_fr else "💰 Depenses",
    "sample_payroll": "👥 Payroll" if not is_fr else "👥 Paie",
    "sample_inventory": "📦 Inventory" if not is_fr else "📦 Inventaire",
    "results": "🎯 AI Detection Results" if not is_fr else "🎯 Resultats IA",
    "download_excel": "📥 Download Full Report" if not is_fr else "📥 Telecharger Rapport",
    "data_preview": "📋 Sample Data Preview" if not is_fr else "📋 Apercu des Donnees",
    "critical": "High Anomaly" if not is_fr else "Anomalie Haute",
    "high": "Medium Anomaly" if not is_fr else "Anomalie Moyenne",
    "medium": "Low Anomaly" if not is_fr else "Anomalie Basse",
    "low": "Normal" if not is_fr else "Normal",
    "normal": "Normal",
    "explanation": "Explanation" if not is_fr else "Explication",
    "level": "Anomaly Level" if not is_fr else "Niveau Anomalie",
    "avg_dev": "Avg Deviation" if not is_fr else "Deviation Moy",
    "anomaly_score": "Anomaly Score" if not is_fr else "Score Anomalie",
    "welcome_title": "👈 Select Sample Data to Start" if not is_fr else "👈 Selectionnez des Donnees",
    "welcome_text": "Choose a sample dataset in the sidebar to see AI in action!" if not is_fr else "Choisissez un jeu de donnees!",
    "get_full": "Get Full Desktop Version" if not is_fr else "Version Complete",
    "full_features": "Isolation Forest AI • Data Aggregation • Custom Ratios • 100% Offline • Unlimited Files" if not is_fr else "IA Isolation Forest • Agregation • Ratios • Hors-ligne",
    "what_is": "What is AynalyxAI?" if not is_fr else "Quest-ce que AynalyxAI?",
    "what_is_desc": "AynalyxAI is an intelligent financial analysis tool that automatically detects anomalies, errors, and irregularities in your accounting data using advanced AI algorithms." if not is_fr else "AynalyxAI detecte automatiquement les anomalies et irregularites dans vos donnees comptables.",
    "how_works": "How It Works" if not is_fr else "Comment ca Marche",
    "step1": "Upload your Excel/CSV data" if not is_fr else "Telechargez vos donnees",
    "step2": "AI analyzes patterns" if not is_fr else "L IA analyse les patterns",
    "step3": "Anomalies flagged with severity" if not is_fr else "Anomalies signalees",
    "step4": "Export color-coded reports" if not is_fr else "Exportez des rapports",
    "adv1_title": "Save Time" if not is_fr else "Gain de Temps",
    "adv1_desc": "Analyze thousands of transactions in seconds" if not is_fr else "Analysez des milliers de transactions",
    "adv2_title": "Reduce Errors" if not is_fr else "Reduire les Erreurs",
    "adv2_desc": "Catch mistakes humans might miss" if not is_fr else "Detectez les erreurs cachees",
    "adv3_title": "Detect Anomalies" if not is_fr else "Detecter les Anomalies",
    "adv3_desc": "Identify suspicious patterns early" if not is_fr else "Identifiez les patterns suspects",
    "example_desc": "Sample data with embedded anomalies. AI will identify unusual patterns." if not is_fr else "Donnees avec anomalies. L IA identifiera les patterns inhabituels.",
    "all_rows_note": "Showing ALL rows sorted by risk. Anomalies highlighted, normal shown for context." if not is_fr else "TOUTES les lignes triees par risque.",
    "aggregation": "📊 Aggregation (Optional)" if not is_fr else "📊 Agregation (Optionnel)",
    "group_by": "Group by column" if not is_fr else "Grouper par colonne",
    "no_grouping": "No grouping" if not is_fr else "Pas de regroupement",
    "aggregated_results": "Aggregated Results" if not is_fr else "Resultats Agreges",
    "run_analysis": "🚀 Run Analysis" if not is_fr else "🚀 Lancer l'Analyse",
}

# Thresholds matching the real app (modeling.py)
THRESHOLD_HIGH = 2.0
THRESHOLD_MEDIUM = 1.2
THRESHOLD_HIGH_AI = -0.15
THRESHOLD_MEDIUM_AI = -0.05

def gen_invoices():
    np.random.seed(42)
    n = 50
    return pd.DataFrame({
        'Invoice_ID': [f'INV-{1000+i}' for i in range(n)],
        'Client': np.random.choice(['Acme Corp', 'TechStart', 'Global Services', 'Local Shop', 'Enterprise'], n),
        'Amount': np.concatenate([np.random.normal(1500, 400, n-6), [18500, 22000, 15, 8, 19800, 12]]),
        'Quantity': np.concatenate([np.random.randint(1, 25, n-6), [180, 250, 1, 1, 200, 1]]),
        'Date': pd.date_range('2024-01-01', periods=n, freq='D').strftime('%Y-%m-%d').tolist()
    })

def gen_expenses():
    np.random.seed(43)
    n = 40
    return pd.DataFrame({
        'Expense_ID': [f'EXP-{2000+i}' for i in range(n)],
        'Category': np.random.choice(['Travel', 'Office', 'Software', 'Marketing', 'Utilities'], n),
        'Vendor': np.random.choice(['Amazon', 'Office Depot', 'Google', 'Airlines', 'Power Co'], n),
        'Amount': np.concatenate([np.random.normal(280, 120, n-5), [5800, 7200, 5, 3, 6500]]),
        'Date': pd.date_range('2024-01-01', periods=n, freq='D').strftime('%Y-%m-%d').tolist()
    })

def gen_payroll():
    np.random.seed(44)
    n = 30
    return pd.DataFrame({
        'Employee_ID': [f'EMP-{100+i}' for i in range(n)],
        'Department': np.random.choice(['Sales', 'Engineering', 'HR', 'Marketing', 'Finance'], n),
        'Salary': np.concatenate([np.random.normal(5200, 800, n-4), [28000, 32000, 450, 380]]),
        'Hours': np.concatenate([np.random.normal(160, 12, n-4), [280, 310, 35, 42]]),
        'Bonus': np.concatenate([np.random.normal(400, 150, n-4), [8500, 12000, 0, 0]])
    })

def gen_inventory():
    np.random.seed(45)
    n = 35
    df = pd.DataFrame({
        'SKU': [f'SKU-{3000+i}' for i in range(n)],
        'Product': np.random.choice(['Widget A', 'Gadget B', 'Tool C', 'Part D', 'Supply E'], n),
        'Quantity': np.concatenate([np.random.randint(80, 400, n-5), [5500, 8200, 2, 1, 6800]]),
        'Unit_Cost': np.concatenate([np.random.normal(28, 8, n-5), [380, 520, 2, 1, 450]])
    })
    df['Total_Value'] = df['Quantity'] * df['Unit_Cost']
    return df

# ============================================================================
# CALCULATION FUNCTIONS - UNTOUCHED
# ============================================================================
def run_anomaly_detection(df, numeric_cols, is_fr):
    """
    Run anomaly detection matching the real app (modeling.py)
    Uses Isolation Forest (negative scores) + Z-score composite
    """
    results = df.copy()
    X = df[numeric_cols].astype(float).copy()
    
    # ISOLATION FOREST (AI Score - NEGATIVE for anomalies)
    iso = IsolationForest(n_estimators=120, contamination=0.05, random_state=42)
    iso.fit(X)
    results['Anomaly_Score'] = np.round(iso.decision_function(X), 2)
    
    # Z-SCORE CALCULATION
    scaler = StandardScaler()
    z_vals = scaler.fit_transform(X)
    
    for i, col in enumerate(numeric_cols):
        results[f'Deviation_{col}'] = np.round(z_vals[:, i], 2)
    
    results['Average_Deviation'] = np.round(np.mean(np.abs(z_vals), axis=1), 2)
    
    def generate_explanation(row):
        explanations = []
        for col in numeric_cols:
            z = row[f'Deviation_{col}']
            if abs(z) >= 1.5:
                direction = "above avg" if z > 0 else "below avg"
                if is_fr:
                    direction = "au-dessus moy" if z > 0 else "en-dessous moy"
                col_clean = col.replace('_', ' ')
                explanations.append(f"{col_clean} = {abs(z):.2f}x {direction}")
        if len(explanations) == 0:
            return "Normal range" if not is_fr else "Plage normale"
        return ", ".join(explanations[:3])
    
    results['Anomaly_Explanation'] = results.apply(generate_explanation, axis=1)
    
    def classify_level(row):
        z_comp = row['Average_Deviation']
        ai_score = row['Anomaly_Score']
        
        both_high = (z_comp >= THRESHOLD_HIGH and ai_score <= THRESHOLD_HIGH_AI)
        stat_extreme = z_comp >= THRESHOLD_HIGH * 1.5
        ai_extreme = ai_score <= THRESHOLD_HIGH_AI * 1.5 and z_comp >= THRESHOLD_MEDIUM
        
        if both_high or stat_extreme or ai_extreme:
            return t['critical']
        elif z_comp >= THRESHOLD_MEDIUM or ai_score <= THRESHOLD_MEDIUM_AI:
            return t['high']
        else:
            return t['low']
    
    results['Anomaly_Level'] = results.apply(classify_level, axis=1)
    
    return results

def run_aggregation(df, group_col, numeric_cols, is_fr):
    """
    Aggregate data by a column and run anomaly detection on aggregated results
    """
    agg_dict = {col: 'sum' for col in numeric_cols}
    agg_df = df.groupby(group_col).agg(agg_dict).reset_index()
    agg_df['Count'] = df.groupby(group_col).size().values
    agg_numeric_cols = numeric_cols + ['Count']
    results = run_anomaly_detection(agg_df, agg_numeric_cols, is_fr)
    return results

# ============================================================================

st.markdown(f'<div class="main-header"><h1>🔬 {t["title"]}</h1><p>{t["subtitle"]}</p></div>', unsafe_allow_html=True)
st.markdown(f'<div class="demo-badge">{t["demo_notice"]}</div>', unsafe_allow_html=True)

# Sidebar - Sample Selection
with st.sidebar:
    st.markdown("---")
    st.markdown(f'<p class="sidebar-title">{t["sample_title"]}</p>', unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        b1 = st.button(t['sample_invoices'], use_container_width=True, type="primary")
        b3 = st.button(t['sample_payroll'], use_container_width=True, type="primary")
    with c2:
        b2 = st.button(t['sample_expenses'], use_container_width=True, type="primary")
        b4 = st.button(t['sample_inventory'], use_container_width=True, type="primary")
    
    if b1: st.session_state['sample_type'] = 'invoices'
    elif b2: st.session_state['sample_type'] = 'expenses'
    elif b3: st.session_state['sample_type'] = 'payroll'
    elif b4: st.session_state['sample_type'] = 'inventory'
    
    st.markdown("---")
    st.markdown(f"**{t['how_works']}**")
    st.markdown(f"1️⃣ {t['step1']}")
    st.markdown(f"2️⃣ {t['step2']}")
    st.markdown(f"3️⃣ {t['step3']}")
    st.markdown(f"4️⃣ {t['step4']}")

# Load data
df, data_name = None, ""
if 'sample_type' in st.session_state:
    s = st.session_state['sample_type']
    if s == 'invoices': df, data_name = gen_invoices(), "Invoices"
    elif s == 'expenses': df, data_name = gen_expenses(), "Expenses"
    elif s == 'payroll': df, data_name = gen_payroll(), "Payroll"
    elif s == 'inventory': df, data_name = gen_inventory(), "Inventory"

if df is not None:
    # Info boxes
    st.markdown(f'<div class="feature-box"><h3>💡 {t["what_is"]}</h3><p>{t["what_is_desc"]}</p></div>', unsafe_allow_html=True)
    st.markdown(f'''<div class="advantage-grid">
        <div class="advantage-item"><div class="icon">⏱️</div><h4>{t['adv1_title']}</h4><p>{t['adv1_desc']}</p></div>
        <div class="advantage-item"><div class="icon">🎯</div><h4>{t['adv2_title']}</h4><p>{t['adv2_desc']}</p></div>
        <div class="advantage-item"><div class="icon">🔍</div><h4>{t['adv3_title']}</h4><p>{t['adv3_desc']}</p></div>
    </div>''', unsafe_allow_html=True)
    
    # Data Preview
    st.subheader(f"{t['data_preview']} — {data_name}")
    st.caption(t['example_desc'])
    st.dataframe(df.head(15), use_container_width=True, hide_index=True)
    st.caption(f"📊 {len(df)} rows × {len(df.columns)} columns")
    
    # Get columns
    numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    categorical_cols = df.select_dtypes(include=['object']).columns.tolist()
    
    # Aggregation Option
    st.markdown("---")
    col_agg, col_btn = st.columns([3, 1])
    with col_agg:
        group_options = [t['no_grouping']] + categorical_cols
        group_col = st.selectbox(f"📊 {t['group_by']}", group_options)
    with col_btn:
        st.markdown("<br>", unsafe_allow_html=True)
        run_btn = st.button(t['run_analysis'], type="primary", use_container_width=True)
    
    use_aggregation = group_col != t['no_grouping']
    
    # Run Analysis
    if run_btn and numeric_cols:
        st.markdown("---")
        st.subheader(t['results'])
        
        with st.spinner("🤖 AI Analyzing..."):
            # Run calculation
            if use_aggregation:
                st.info(f"📊 {t['aggregated_results']}: **{group_col}**")
                results_sorted = run_aggregation(df, group_col, numeric_cols, is_fr)
            else:
                results_sorted = run_anomaly_detection(df, numeric_cols, is_fr)
            
            # Sort by Anomaly_Score (most negative first = most anomalous)
            results_sorted = results_sorted.sort_values('Anomaly_Score', ascending=True).reset_index(drop=True)
            
            # Count anomalies
            n_high = len(results_sorted[results_sorted['Anomaly_Level'] == t['critical']])
            n_med = len(results_sorted[results_sorted['Anomaly_Level'] == t['high']])
            n_normal = len(results_sorted[results_sorted['Anomaly_Level'] == t['low']])
            n_total = n_high + n_med
            
            # Stats Cards
            c1, c2, c3, c4 = st.columns(4)
            c1.markdown(f'<div class="stat-card stat-critical"><div style="font-size:2rem;font-weight:bold;color:#dc2626;">{n_high}</div><div>🔴 High Anomaly</div></div>', unsafe_allow_html=True)
            c2.markdown(f'<div class="stat-card stat-medium"><div style="font-size:2rem;font-weight:bold;color:#f59e0b;">{n_med}</div><div>🟡 Medium Anomaly</div></div>', unsafe_allow_html=True)
            c3.markdown(f'<div class="stat-card stat-low"><div style="font-size:2rem;font-weight:bold;color:#22c55e;">{n_normal}</div><div>🟢 Normal</div></div>', unsafe_allow_html=True)
            c4.markdown(f'<div class="stat-card" style="border-color:#667eea;"><div style="font-size:2rem;font-weight:bold;color:#667eea;">{n_total}</div><div>📊 Total Flagged</div></div>', unsafe_allow_html=True)
            
            st.markdown("<br>", unsafe_allow_html=True)
            st.success(f"✅ **{len(results_sorted)}** rows analyzed — **{n_total}** anomalies detected — **{n_normal}** normal")
            st.info(f"ℹ️ {t['all_rows_note']}")
            
            # Prepare display dataframe - keep original columns + analysis columns
            # Remove individual Deviation_ columns for cleaner view
            display_cols = [c for c in results_sorted.columns if not c.startswith('Deviation_')]
            display_df = results_sorted[display_cols].copy()
            
            # Rename analysis columns
            rename_map = {
                'Anomaly_Level': t['level'], 
                'Average_Deviation': t['avg_dev'],
                'Anomaly_Score': t['anomaly_score'],
                'Anomaly_Explanation': t['explanation']
            }
            display_df = display_df.rename(columns=rename_map)
            
            # Styling functions
            def color_level(val):
                if val == t['critical']: return 'background-color:#dc2626;color:white;font-weight:bold;'
                if val == t['high']: return 'background-color:#f59e0b;color:white;font-weight:bold;'
                if val == t['low']: return 'background-color:#22c55e;color:white;font-weight:bold;'
                return 'background-color:#f1f5f9;color:#64748b;'
            
            def color_row(row):
                lv = row[t['level']]
                if lv == t['critical']: return ['background-color:#fee2e2;'] * len(row)
                if lv == t['high']: return ['background-color:#fffbeb;'] * len(row)
                return [''] * len(row)
            
            # Display styled dataframe
            styled = display_df.style.apply(color_row, axis=1).map(color_level, subset=[t['level']])
            st.dataframe(styled, use_container_width=True, height=500)
            
            # Excel Export
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                display_df.to_excel(writer, index=False, sheet_name='Results')
                ws = writer.sheets['Results']
                fill_h = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
                fill_c = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')
                fill_cr = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
                fill_m = PatternFill(start_color='F59E0B', end_color='F59E0B', fill_type='solid')
                fill_mr = PatternFill(start_color='FFFBEB', end_color='FFFBEB', fill_type='solid')
                fill_l = PatternFill(start_color='22C55E', end_color='22C55E', fill_type='solid')
                font_w = Font(color='FFFFFF', bold=True)
                bdr = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                for cell in ws[1]: cell.fill, cell.font, cell.border = fill_h, font_w, bdr
                lv_idx = list(display_df.columns).index(t['level']) + 1
                for r in range(2, len(display_df) + 2):
                    lv = ws.cell(row=r, column=lv_idx).value
                    rf, lf = None, None
                    if lv == t['critical']: rf, lf = fill_cr, fill_c
                    elif lv == t['high']: rf, lf = fill_mr, fill_m
                    elif lv == t['low']: rf, lf = None, fill_l
                    for c in range(1, len(display_df.columns) + 1):
                        cell = ws.cell(row=r, column=c)
                        cell.border = bdr
                        if rf: cell.fill = rf
                        if c == lv_idx and lf: cell.fill, cell.font = lf, font_w
                for i, col in enumerate(display_df.columns, 1):
                    ws.column_dimensions[ws.cell(1, i).column_letter].width = min(45, max(len(str(col)), 12) + 2)
                ws.auto_filter.ref = f"A1:{ws.cell(1, len(display_df.columns)).column_letter}{len(display_df) + 1}"
                ws.freeze_panes = 'A2'
            output.seek(0)
            
            st.download_button(t['download_excel'], output, f"aynalyxai_{data_name.lower()}_report.xlsx", 
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")

        # CTA Box
        st.markdown(f'<div class="cta-box"><h3>🚀 {t["get_full"]}</h3><p style="color:#555;">{t["full_features"]}</p><a href="https://aynalyx.gumroad.com/l/qpqmv" target="_blank" class="cta-button">💎 Get AynalyxAI Pro</a></div>', unsafe_allow_html=True)

else:
    # Welcome screen
    st.markdown(f'<div class="feature-box"><h3>💡 {t["what_is"]}</h3><p>{t["what_is_desc"]}</p></div>', unsafe_allow_html=True)
    st.markdown(f'''<div class="advantage-grid">
        <div class="advantage-item"><div class="icon">⏱️</div><h4>{t['adv1_title']}</h4><p>{t['adv1_desc']}</p></div>
        <div class="advantage-item"><div class="icon">🎯</div><h4>{t['adv2_title']}</h4><p>{t['adv2_desc']}</p></div>
        <div class="advantage-item"><div class="icon">🔍</div><h4>{t['adv3_title']}</h4><p>{t['adv3_desc']}</p></div>
    </div>''', unsafe_allow_html=True)
    st.markdown(f'<div class="how-it-works"><h4>📋 {t["how_works"]}</h4><ul><li>1️⃣ {t["step1"]}</li><li>2️⃣ {t["step2"]}</li><li>3️⃣ {t["step3"]}</li><li>4️⃣ {t["step4"]}</li></ul></div>', unsafe_allow_html=True)
    st.markdown(f'<div style="text-align:center;padding:3rem 2rem;background:linear-gradient(135deg,#f8fafc 0%,#e2e8f0 100%);border-radius:16px;margin-top:1rem;"><div style="font-size:4rem;margin-bottom:1rem;">👈</div><h2 style="color:#1e293b;">{t["welcome_title"]}</h2><p style="color:#64748b;font-size:1.1rem;">{t["welcome_text"]}</p></div>', unsafe_allow_html=True)

st.markdown("---")
st.markdown('<div style="text-align:center;color:#94a3b8;padding:1rem;">© 2025 Mubsira Analytics | <a href="https://aynalyx.gumroad.com/l/qpqmv" style="color:#667eea;">Get Full Version</a></div>', unsafe_allow_html=True)
